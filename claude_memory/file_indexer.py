"""
File indexing system for Claude Memory.
Discovers and indexes documents on disk for universal search.
"""

import os
import sqlite3
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Set, Tuple
import mimetypes

from .database import get_connection


# Supported file extensions
SUPPORTED_EXTENSIONS = {
    # Text files
    '.txt', '.md', '.markdown', '.log', '.csv', '.json', '.xml', '.yaml', '.yml',
    '.ini', '.cfg', '.conf', '.config',
    # Documents
    '.docx', '.doc', '.odt',
    # Spreadsheets
    '.xlsx', '.xls', '.ods',
    # PDFs
    '.pdf',
    # Code files
    '.py', '.js', '.java', '.cpp', '.c', '.h', '.cs', '.php', '.rb', '.go',
    '.html', '.css', '.sql', '.sh', '.bat', '.ps1'
}

# Directories to skip (system, temporary, caches)
SKIP_DIRECTORIES = {
    '$RECYCLE.BIN', 'System Volume Information', 'Recovery', 'ProgramData',
    'Windows', 'Program Files', 'Program Files (x86)', 'AppData',
    'node_modules', '__pycache__', '.git', '.svn', '.hg',
    'venv', 'env', '.venv', 'site-packages',
    'cache', 'Cache', 'Temp', 'temp', 'tmp'
}

# Minimum file size (10 bytes) and maximum (100 MB)
MIN_FILE_SIZE = 10
MAX_FILE_SIZE = 100 * 1024 * 1024


def is_text_file(file_path: Path) -> bool:
    """Check if a file is a text-based file."""
    # Check extension first
    if file_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        return False

    # Check file size
    try:
        size = file_path.stat().st_size
        if size < MIN_FILE_SIZE or size > MAX_FILE_SIZE:
            return False
    except:
        return False

    return True


def should_skip_directory(dir_name: str) -> bool:
    """Check if a directory should be skipped."""
    return dir_name in SKIP_DIRECTORIES or dir_name.startswith('.')


def extract_text_preview(file_path: Path, max_chars: int = 1000) -> str:
    """
    Extract preview text from a file.
    Supports plain text, Word docs, Excel sheets, and PDFs.
    """
    try:
        ext = file_path.suffix.lower()

        # Plain text files
        if ext in {'.txt', '.md', '.log', '.csv', '.json', '.xml', '.yaml', '.yml',
                   '.py', '.js', '.java', '.cpp', '.c', '.h', '.cs', '.html', '.css', '.sql'}:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read(max_chars)
                    return content[:max_chars]
            except UnicodeDecodeError:
                # Try with different encoding
                try:
                    with open(file_path, 'r', encoding='latin-1') as f:
                        content = f.read(max_chars)
                        return content[:max_chars]
                except:
                    return "[Binary or unreadable file]"

        # Word documents
        elif ext == '.docx':
            try:
                from docx import Document
                doc = Document(file_path)
                text = '\n'.join([para.text for para in doc.paragraphs])
                return text[:max_chars] if text else "[Empty Word document]"
            except Exception as e:
                return f"[Error reading Word doc: {str(e)[:50]}]"

        # Excel spreadsheets
        elif ext in {'.xlsx', '.xls'}:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, read_only=True, data_only=True)
                text_parts = []

                # Read first sheet only for preview
                ws = wb.active
                for row in ws.iter_rows(max_row=20, values_only=True):
                    row_text = ' | '.join([str(cell) if cell is not None else '' for cell in row])
                    if row_text.strip():
                        text_parts.append(row_text)

                wb.close()
                text = '\n'.join(text_parts)
                return text[:max_chars] if text else "[Empty Excel sheet]"
            except Exception as e:
                return f"[Error reading Excel: {str(e)[:50]}]"

        # PDF documents
        elif ext == '.pdf':
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(file_path)
                text_parts = []

                # Extract text from first 3 pages for preview
                for page_num in range(min(3, len(doc))):
                    page = doc[page_num]
                    text_parts.append(page.get_text())

                doc.close()
                text = '\n'.join(text_parts)
                return text[:max_chars] if text else "[Empty or image-only PDF]"
            except Exception as e:
                return f"[Error reading PDF: {str(e)[:50]}]"
        else:
            return "[Supported file type]"

    except Exception as e:
        return f"[Error reading file: {str(e)}]"


def scan_directory(directory: Path, progress_callback=None) -> Dict[str, any]:
    """
    Recursively scan a directory for supported files.

    Returns:
        Dictionary with scan results including file count and list of files
    """
    results = {
        'total_files': 0,
        'by_type': {},
        'files': []
    }

    try:
        for root, dirs, files in os.walk(directory):
            # Skip certain directories
            dirs[:] = [d for d in dirs if not should_skip_directory(d)]

            root_path = Path(root)

            for filename in files:
                file_path = root_path / filename

                if is_text_file(file_path):
                    try:
                        stat = file_path.stat()
                        ext = file_path.suffix.lower()

                        file_info = {
                            'path': str(file_path),
                            'name': filename,
                            'type': ext,
                            'size': stat.st_size,
                            'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                        }

                        results['files'].append(file_info)
                        results['total_files'] += 1

                        # Track by file type
                        if ext not in results['by_type']:
                            results['by_type'][ext] = 0
                        results['by_type'][ext] += 1

                        if progress_callback:
                            progress_callback(results['total_files'])

                    except Exception as e:
                        # Skip files we can't access
                        continue

    except Exception as e:
        print(f"Error scanning directory {directory}: {e}")

    return results


def add_watched_folder(folder_path: str, is_monitored: bool = False) -> int:
    """
    Add a folder to the watched folders list.

    Returns:
        Folder ID
    """
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT OR REPLACE INTO watched_folders (path, is_monitored, added_date)
        VALUES (?, ?, ?)
    """, (folder_path, 1 if is_monitored else 0, datetime.now().isoformat()))

    folder_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return folder_id


def index_files(folder_id: int, files: List[Dict], progress_callback=None) -> int:
    """
    Index a list of files into the database.

    Returns:
        Number of files indexed
    """
    conn = get_connection()
    cursor = conn.cursor()

    indexed_count = 0

    for i, file_info in enumerate(files):
        try:
            # Extract preview
            preview = extract_text_preview(Path(file_info['path']))

            # Insert or replace file record
            cursor.execute("""
                INSERT OR REPLACE INTO indexed_files
                (file_path, file_name, file_type, file_size, modified_date, content_preview, folder_id)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                file_info['path'],
                file_info['name'],
                file_info['type'],
                file_info['size'],
                file_info['modified'],
                preview,
                folder_id
            ))

            indexed_count += 1

            if progress_callback and i % 10 == 0:
                progress_callback(i + 1, len(files))

        except Exception as e:
            print(f"Error indexing {file_info['path']}: {e}")
            continue

    # Update folder file count and last scan date
    cursor.execute("""
        UPDATE watched_folders
        SET file_count = ?, last_scan_date = ?
        WHERE id = ?
    """, (indexed_count, datetime.now().isoformat(), folder_id))

    conn.commit()
    conn.close()

    return indexed_count


def get_watched_folders() -> List[Dict]:
    """Get all watched folders."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, path, is_monitored, last_scan_date, file_count, enabled
        FROM watched_folders
        ORDER BY path
    """)

    folders = []
    for row in cursor.fetchall():
        folders.append({
            'id': row[0],
            'path': row[1],
            'is_monitored': bool(row[2]),
            'last_scan_date': row[3],
            'file_count': row[4],
            'enabled': bool(row[5])
        })

    conn.close()
    return folders


def search_indexed_files(query: str, file_type: Optional[str] = None) -> List[Dict]:
    """
    Search indexed files using FTS5.

    Args:
        query: Search query
        file_type: Optional file extension filter (e.g., '.txt')

    Returns:
        List of matching files
    """
    if not query:
        return []

    conn = get_connection()
    cursor = conn.cursor()

    if file_type:
        cursor.execute("""
            SELECT f.id, f.file_path, f.file_name, f.file_type, f.file_size,
                   f.modified_date, f.content_preview
            FROM indexed_files f
            JOIN files_fts fts ON f.id = fts.rowid
            WHERE files_fts MATCH ? AND f.file_type = ?
            ORDER BY rank
            LIMIT 50
        """, (query, file_type))
    else:
        cursor.execute("""
            SELECT f.id, f.file_path, f.file_name, f.file_type, f.file_size,
                   f.modified_date, f.content_preview
            FROM indexed_files f
            JOIN files_fts fts ON f.id = fts.rowid
            WHERE files_fts MATCH ?
            ORDER BY rank
            LIMIT 50
        """, (query,))

    results = []
    for row in cursor.fetchall():
        results.append({
            'id': row[0],
            'file_path': row[1],
            'file_name': row[2],
            'file_type': row[3],
            'file_size': row[4],
            'modified_date': row[5],
            'content_preview': row[6]
        })

    conn.close()
    return results


def remove_watched_folder(folder_id: int) -> None:
    """Remove a watched folder and all its indexed files."""
    conn = get_connection()
    cursor = conn.cursor()

    # Delete indexed files (CASCADE will handle this)
    cursor.execute("DELETE FROM watched_folders WHERE id = ?", (folder_id,))

    conn.commit()
    conn.close()


def get_file_type_icon(file_type: str) -> str:
    """Get an emoji icon for a file type."""
    icons = {
        '.txt': '📄', '.md': '📝', '.log': '📋',
        '.docx': '📘', '.doc': '📘',
        '.xlsx': '📊', '.xls': '📊', '.csv': '📊',
        '.pdf': '📕',
        '.py': '🐍', '.js': '📜', '.html': '🌐',
        '.json': '{}', '.xml': '📋', '.yaml': '⚙️',
    }
    return icons.get(file_type.lower(), '📄')
